import openpyxl
import re

# Function to process and clean each cell
def clean_cell_value(value):
    if isinstance(value, str):

        value = re.sub(r'[\r\n]+', '', value)
        value = re.sub(r'\s+', ' ', value).strip() 
    return value

# Storing the file path in a variable for easy recollection
file_path = 'sample.xlsx'
wb = openpyxl.load_workbook(file_path)


for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"Cleaning data in sheet: {sheet}")

    # Loop through each row and column in the sheet
    for row in ws.iter_rows():
        for cell in row:
            # Calling the function to clean the cell
            cell.value = clean_cell_value(cell.value)

output_file_path = 'Sample (Clean).xlsx'
wb.save(output_file_path)

print(f"Cleaning complete. The newfile has been saved as: {output_file_path}")
